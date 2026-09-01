# -*- coding: utf-8 -*-
"""
Інтеграція з OneDrive/SharePoint через Microsoft Graph (app-only, client
credentials). Підтримує КІЛЬКА проєктів одночасно (config.PROJECTS) —
кожен проєкт має свою окрему папку на OneDrive, свою структуру дат і
свій окремий dashboard.xlsx у корені своєї папки.

Потрібні змінні оточення (див. .env.example):
    MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET, PROJECTS_JSON (або MS_SHARE_URL)

Логіка дат: перед кожним завантаженням бот кладе файл у підпапку з
сьогоднішньою датою (YYYY-MM-DD) всередині папки ПРОЄКТУ цього рейсу.
Якщо такої підпапки ще немає — Graph API створює її автоматично.

Якщо OneDrive ще не налаштований — ONEDRIVE_ENABLED=False і всі функції
тихо пропускають завантаження (файли лишаються тільки локально в data/).
"""
import os
import base64
import logging
import time
from datetime import datetime

import msal
import requests
from openpyxl import Workbook, load_workbook

import config
import database as db

log = logging.getLogger(__name__)

GRAPH_ROOT = "https://graph.microsoft.com/v1.0"
DASHBOARD_HEADERS = [
    "Рейс №", "Водій", "Телефон", "Тягач", "Причіп", "Вантаж", "Маршрут",
    "Початок", "Завершення", "Статус", "Паузи", "Дод. зупинки", "ЧП", "Звіт (OneDrive)",
]

# Резолвимо посилання на папку кожного проєкту один раз і кешуємо
# (drive_id, item_id) в пам'яті процесу, окремо на кожен project_key.
_project_folder_cache: dict[str, tuple[str, str]] = {}


def _get_token():
    app = msal.ConfidentialClientApplication(
        client_id=config.MS_CLIENT_ID,
        client_credential=config.MS_CLIENT_SECRET,
        authority=f"https://login.microsoftonline.com/{config.MS_TENANT_ID}",
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Не вдалось отримати токен Graph API: {result.get('error_description')}")
    return result["access_token"]


def _encode_share_url(url: str) -> str:
    """Кодує звичайне посилання 'Поділитися' у формат, який приймає /shares/{id}."""
    b64 = base64.urlsafe_b64encode(url.encode("utf-8")).decode("utf-8")
    b64 = b64.rstrip("=")
    return "u!" + b64


def _resolve_project_folder(headers, project_key: str):
    """Повертає (drive_id, item_id) папки конкретного проєкту. Кешується в пам'яті."""
    if project_key in _project_folder_cache:
        return _project_folder_cache[project_key]

    project = config.PROJECTS_BY_KEY.get(project_key)
    if not project:
        raise RuntimeError(f"Проєкт '{project_key}' не знайдено в конфігурації (config.PROJECTS)")

    share_id = _encode_share_url(project["share_url"])
    url = f"{GRAPH_ROOT}/shares/{share_id}/driveItem"
    resp = requests.get(url, headers=headers, timeout=30)
    resp.raise_for_status()
    item = resp.json()
    drive_id = item["parentReference"]["driveId"]
    item_id = item["id"]
    _project_folder_cache[project_key] = (drive_id, item_id)
    log.info("OneDrive: папку проєкту '%s' розпізнано (drive=%s, item=%s)", project_key, drive_id, item_id)
    return _project_folder_cache[project_key]


def _today_folder_name() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def upload_file(local_path: str, remote_subpath: str, project_key: str) -> str | None:
    """
    Завантажує файл у {папка проєкту}/{remote_subpath} на OneDrive.
    Проміжні папки (в т.ч. папка з датою) створюються автоматично.
    Повертає webUrl файлу, або None якщо OneDrive/проєкт не налаштовані чи
    стався збій (файл лишається доступним локально — рейс не втрачається).

    На тимчасові помилки (423 Locked, 429, 503) робимо кілька повторних
    спроб із паузою, а не здаємось одразу.
    """
    if not config.ONEDRIVE_ENABLED:
        log.info("OneDrive не налаштований — файл %s лишається тільки локально", local_path)
        return None
    if not project_key or project_key not in config.PROJECTS_BY_KEY:
        log.warning(
            "OneDrive: невідомий або відсутній проєкт ('%s') для %s — файл лишається тільки локально",
            project_key, local_path,
        )
        return None

    RETRYABLE_STATUSES = {423, 429, 503}
    MAX_ATTEMPTS = 4
    DELAYS = [3, 8, 20]  # секунди між спробами (зростаюча пауза)

    with open(local_path, "rb") as f:
        data = f.read()
    remote_subpath = remote_subpath.replace("//", "/").lstrip("/")

    last_error = None
    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            token = _get_token()
            headers = {"Authorization": f"Bearer {token}"}
            drive_id, item_id = _resolve_project_folder(headers, project_key)

            url = f"{GRAPH_ROOT}/drives/{drive_id}/items/{item_id}:/{remote_subpath}:/content"
            resp = requests.put(url, headers=headers, data=data, timeout=60)

            if resp.status_code in RETRYABLE_STATUSES and attempt < MAX_ATTEMPTS:
                delay = DELAYS[attempt - 1]
                log.warning(
                    "OneDrive: тимчасова помилка %s для [%s] %s (спроба %d/%d), повтор через %ds",
                    resp.status_code, project_key, remote_subpath, attempt, MAX_ATTEMPTS, delay,
                )
                time.sleep(delay)
                continue

            resp.raise_for_status()
            return resp.json().get("webUrl")
        except Exception as e:
            last_error = e
            if attempt < MAX_ATTEMPTS:
                delay = DELAYS[attempt - 1]
                log.warning(
                    "OneDrive: помилка завантаження [%s] %s (спроба %d/%d): %s — повтор через %ds",
                    project_key, remote_subpath, attempt, MAX_ATTEMPTS, e, delay,
                )
                time.sleep(delay)

    log.error(
        "Помилка завантаження на OneDrive [%s] (%s) після %d спроб: %s",
        project_key, local_path, MAX_ATTEMPTS, last_error,
    )
    return None


def _download_file(remote_subpath: str, project_key: str, local_dest_path: str) -> bool:
    """Тягне файл з OneDrive проєкту на локальний шлях. Повертає True, якщо
    вдалось (файл існує і завантажений), False якщо файлу немає або стався збій."""
    if not config.ONEDRIVE_ENABLED or not project_key or project_key not in config.PROJECTS_BY_KEY:
        return False
    try:
        token = _get_token()
        headers = {"Authorization": f"Bearer {token}"}
        drive_id, item_id = _resolve_project_folder(headers, project_key)
        remote_subpath = remote_subpath.replace("//", "/").lstrip("/")
        url = f"{GRAPH_ROOT}/drives/{drive_id}/items/{item_id}:/{remote_subpath}:/content"
        resp = requests.get(url, headers=headers, timeout=60)
        if resp.status_code == 404:
            return False
        resp.raise_for_status()
        with open(local_dest_path, "wb") as f:
            f.write(resp.content)
        return True
    except Exception as e:
        log.warning("Не вдалось завантажити %s з OneDrive [%s]: %s", remote_subpath, project_key, e)
        return False


def _ensure_local_dashboard(path: str, project_key: str):
    if os.path.exists(path):
        return
    # Локального файлу немає (перший запуск, скинута база для тестів, чи
    # взагалі втрачений диск) — ПЕРШ НІЖ створювати порожній файл з нуля,
    # пробуємо стягнути актуальну версію з OneDrive. Без цього кроку
    # наступне збереження перезаписало б реальну таблицю на OneDrive
    # порожньою, стерши всі попередні рейси, яких немає локально.
    if _download_file("dashboard.xlsx", project_key, path):
        log.info("OneDrive: локального dashboard.xlsx не було — відновив з хмари (проєкт %s)", project_key)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Рейси"
    ws.append(DASHBOARD_HEADERS)
    wb.save(path)


def update_dashboard_row(trip_id: int, onedrive_report_url: str | None):
    """Додає/оновлює рядок зведеної таблиці по рейсу і синхронізує з OneDrive.
    Кожен проєкт має СВІЙ dashboard.xlsx у корені СВОЄЇ папки (не в
    підпапці з датою) — окремий зведений список по кожному проєкту."""
    trip = db.get_trip(trip_id)
    project_key = trip["project_key"]
    dash_path = config.dashboard_path(project_key)
    _ensure_local_dashboard(dash_path, project_key)

    driver = db.get_driver(trip["driver_id"])
    incidents = db.get_incidents(trip_id)
    pauses = db.get_optional_events(trip_id, "pause")
    extra_stops = db.get_optional_events(trip_id, "extra_stop")

    wb = load_workbook(dash_path)
    ws = wb["Рейси"]

    target_row = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value == trip_id:
            target_row = row[0].row
            break

    values = [
        trip_id,
        driver["full_name"] if driver else "",
        driver["phone"] if driver and driver["phone"] else "",
        trip["tractor_number"] or "",
        trip["trailer_number"] or "",
        trip["cargo_description"] or "",
        trip["route"] or "",
        trip["started_at"] or "",
        trip["finished_at"] or "",
        "Завершено" if trip["status"] == "done" else trip["status"],
        len(pauses),
        len(extra_stops),
        len(incidents),
        onedrive_report_url or "",
    ]

    if target_row:
        for col, val in enumerate(values, start=1):
            ws.cell(row=target_row, column=col, value=val)
    else:
        ws.append(values)

    wb.save(dash_path)
    if not project_key:
        return None
    return upload_file(dash_path, "dashboard.xlsx", project_key)


def upload_trip_report(trip_id: int, report_path: str) -> str | None:
    """Кладе звіт у папку проєкту, підпапка сьогоднішньої дати: {дата}/reports/reys_N.docx"""
    trip = db.get_trip(trip_id)
    fname = os.path.basename(report_path)
    return upload_file(report_path, f"{_today_folder_name()}/reports/{fname}", trip["project_key"])


def upload_trip_photo(trip_id: int, photo_path: str) -> str | None:
    """Кладе фото у папку проєкту, підпапка сьогоднішньої дати: {дата}/photos/reys_N/файл.jpg"""
    trip = db.get_trip(trip_id)
    fname = os.path.basename(photo_path)
    return upload_file(photo_path, f"{_today_folder_name()}/photos/reys_{trip_id}/{fname}", trip["project_key"])
